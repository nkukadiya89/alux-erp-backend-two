from datetime import datetime
from io import BytesIO
import os
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import Paragraph

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


class ExportUtility:

    @staticmethod
    def export_excel(
        queryset,
        columns,
        filename="export.xlsx",
        sheet_name="Data",
    ):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header
        for col_num, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Data
        for sr_no, row in enumerate(
            queryset.iterator(chunk_size=1000),
            1,
        ):
            for col_num, (_, field) in enumerate(columns, 1):
                from datetime import datetime
                if field == "sr_no":
                    value = sr_no
                else:
                    value = row.get(field)

                if isinstance(value, datetime):
                   value = value.strftime("%H:%M:%S")

                ws.cell(
                    row=sr_no + 1,
                    column=col_num,
                    value=str(value) if value else "",
                )

                ws.cell(row=sr_no + 1, column=col_num).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        response = HttpResponse(
            output,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        return response

    @staticmethod
    def export_to_pdf(
        queryset, fields, headers, file_name="data.pdf", title="Report", col_widths=None
    ):
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f"attachment; filename={file_name}"

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=50,
            bottomMargin=30,
        )

        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
        elements.append(Spacer(1, 10))

        data = [["Sr. No."] + headers]

        for index, obj in enumerate(queryset, start=1):
            row = [index]
            for field in fields:
                value = getattr(obj, field, "")
                if hasattr(value, "strftime"):
                    value = value.strftime("%d-%m-%Y")
                if value is None:
                    value = ""
                row.append(Paragraph(str(value), styles["Normal"]))
            data.append(row)

        if not col_widths:
            total_width = 800
            col_width = total_width / (len(headers) + 1)
            col_widths = [col_width] * (len(headers) + 1)
        else:
            col_widths = [25] + col_widths

        table = Table(data, colWidths=col_widths, repeatRows=1)

        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.black),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("WORDWRAP", (0, 0), (-1, -1), True),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )

        elements.append(table)

        def add_header_footer(canvas, doc):
            canvas.saveState()

            footer_text = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
            canvas.setFont("Helvetica", 8)
            canvas.drawString(30, 20, footer_text)
            canvas.drawRightString(800, 20, f"Page {doc.page}")

            logo_path = os.path.join("media", "alux_logo.png")
            if os.path.exists(logo_path):
                canvas.drawImage(
                    logo_path,
                    700,
                    540,
                    width=80,
                    height=40,
                    preserveAspectRatio=True,
                    mask="auto",
                )

            canvas.restoreState()

        doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)

        return response
